# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import hashlib
import importlib.util
import json
import os
import re
import sys
import time
from dataclasses import asdict, dataclass
from datetime import datetime
from html import escape
from pathlib import Path
from typing import Any, Mapping

from openpyxl import load_workbook

from business_card_automation import load_automation_settings, run_automation_cycle
from business_card_storage import OperationLock, atomic_save_json, load_json as storage_load_json
from business_card_sending import send_ready_drafts_safely, validate_draft_for_send
from business_card_dashboard import (
    render_dashboard_html as render_dashboard_page,
    summarize_import_history,
    summarize_send_history,
)

try:
    from dotenv import load_dotenv
except ImportError:
    def load_dotenv(dotenv_path: Path | str, override: bool = False) -> bool:
        path = Path(dotenv_path)
        if not path.exists():
            return False
        loaded = False
        for raw_line in path.read_text(encoding="utf-8").splitlines():
            line = raw_line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            if not key:
                continue
            if override or key not in os.environ:
                os.environ[key] = value
                loaded = True
        return loaded


PROJECT_DIR = Path(__file__).resolve().parent
INPUT_DIR = PROJECT_DIR / "inbox"
AUTOMATION_SETTINGS_PATH = INPUT_DIR / "automation_settings.json"
REQUEST_DIR = INPUT_DIR / "requests"
OUTPUT_DIR = PROJECT_DIR / "output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

MAPPING_PATH = INPUT_DIR / "request_sheet_mapping.json"
TEMPLATE_PATH = INPUT_DIR / "vendor_mail_template.json"
DRAFTS_PATH = OUTPUT_DIR / "business_card_drafts.json"
PREVIEW_PATH = OUTPUT_DIR / "business_card_preview.html"
DASHBOARD_PATH = OUTPUT_DIR / "business_card_dashboard.html"
MAIL_FETCH_RESULT_PATH = OUTPUT_DIR / "business_card_mail_fetch_result.json"
SEND_RESULT_PATH = OUTPUT_DIR / "business_card_send_result.json"
STATE_PATH = OUTPUT_DIR / "processed_state.json"
OPERATION_LOCK_PATH = OUTPUT_DIR / "business_card_operation.lock"

load_dotenv(PROJECT_DIR / ".env", override=False)
load_dotenv(PROJECT_DIR / ".env.txt", override=False)

DEFAULT_MAPPING = {
    "header_row_candidates": [1, 2, 3, 4, 5],
    "minimum_header_matches": 4,
    "required_header_fields": ["employee_name"],
    "required_fields": [
        "employee_name",
        "department",
        "title",
        "english_name",
        "mobile",
        "email",
    ],
    "field_aliases": {
        "company_name": ["사업장", "회사명", "법인명", "company", "company name"],
        "department": ["부서", "소속", "department", "dept"],
        "team": ["팀", "team"],
        "title": ["직위", "직급", "직책", "title", "position"],
        "mobile": ["mobile", "휴대폰", "휴대전화", "핸드폰", "cell phone", "phone"],
        "office_phone": ["tel(direct)", "tel", "telephone", "office phone", "사무실 전화", "내선"],
        "employee_name": ["이름", "성명", "신청자명", "name", "employee name"],
        "english_name": ["영문이름", "영문명", "영문 이름", "english name", "name(english)", "eng name"],
        "english_title": ["영문직위", "english title", "eng title", "position(eng)"],
        "english_department": ["영문부서", "english department", "eng department", "department(eng)"],
        "email": ["메일주소", "이메일", "메일", "email", "e-mail"],
        "fax_no": ["fax번호", "fax", "fax no", "fax number"],
        "remarks": ["기타 수정/요청사항", "기타수정요청사항", "요청사항", "비고", "remark", "remarks", "note"],
    },
}

DEFAULT_TEMPLATE = {
    "vendor_to": "",
    "vendor_cc": "",
    "subject_template": "[명함 발주] {{request_date}} / {{request_count}}명",
    "intro_lines": [
        "안녕하세요.",
        "첨부드린 명함신청서(Form) 기준으로 명함 제작 요청드립니다.",
    ],
    "closing_lines": [
        "신청서는 원본 그대로 첨부드립니다.",
        "",
        "감사합니다.",
    ],
    "field_labels": {
        "request_date": "신청일",
        "request_count": "신청 인원",
        "employee_names": "신청자",
        "company_name": "사업장",
        "employee_name": "이름",
        "english_name": "영문이름",
        "department": "부서",
        "team": "팀",
        "title": "직위",
        "mobile": "Mobile",
        "office_phone": "TEL(direct)",
        "english_title": "영문직위",
        "english_department": "영문부서",
        "email": "메일주소",
        "fax_no": "FAX번호",
        "remarks": "기타 수정/요청사항",
    },
    "attach_original_excel": True,
}
DETAIL_FIELD_ORDER = [
    "company_name",
    "department",
    "title",
    "mobile",
    "office_phone",
    "employee_name",
    "english_name",
    "english_title",
    "english_department",
    "email",
    "fax_no",
    "remarks",
]

_ARCHIVE_MODULE: Any | None = None


@dataclass(slots=True)
class DraftRecord:
    draft_id: str
    source_file: str
    source_sheet: str
    source_hash: str
    request_date: str
    request_count: int
    status: str
    missing_fields: list[str]
    subject: str
    vendor_to: str
    vendor_cc: str
    html_body: str
    attachment_paths: list[str]
    requests: list[dict[str, Any]]


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return re.sub(r"\s+", " ", text)


def normalize_header(value: Any) -> str:
    text = normalize_text(value).lower()
    return re.sub(r"[\s_\-:/().\[\]]+", "", text)


def load_json(path: Path, fallback: Any) -> Any:
    return storage_load_json(path, fallback)


def save_json(path: Path, payload: Any) -> None:
    atomic_save_json(path, payload)


def file_sha1(path: Path) -> str:
    sha1 = hashlib.sha1()
    with path.open("rb") as handle:
        while True:
            chunk = handle.read(1024 * 1024)
            if not chunk:
                break
            sha1.update(chunk)
    return sha1.hexdigest()


def render_template_text(template: str, context: dict[str, str]) -> str:
    def replace(match: re.Match[str]) -> str:
        key = match.group(1).strip()
        return context.get(key, "")

    return re.sub(r"\{\{\s*([^{}]+?)\s*\}\}", replace, template)


def get_vendor_recipient(template: dict[str, Any]) -> tuple[str, str]:
    vendor_to_raw = normalize_text(os.getenv("BUSINESS_CARD_VENDOR_TO") or template.get("vendor_to"))
    vendor_cc_raw = normalize_text(os.getenv("BUSINESS_CARD_VENDOR_CC") or template.get("vendor_cc"))

    def split_addresses(raw: str) -> list[str]:
        return [item.strip() for item in re.split(r"[;,\n]", raw) if normalize_text(item)]

    seen: set[str] = set()
    vendor_to_list: list[str] = []
    vendor_cc_list: list[str] = []

    for address in split_addresses(vendor_to_raw):
        key = address.lower()
        if key in seen:
            continue
        seen.add(key)
        vendor_to_list.append(address)

    for address in split_addresses(vendor_cc_raw):
        key = address.lower()
        if key in seen:
            continue
        seen.add(key)
        vendor_cc_list.append(address)

    return "; ".join(vendor_to_list), "; ".join(vendor_cc_list)


def display_cell(value: Any) -> str:
    text = normalize_text(value)
    return text if text else "-"


def detail_fields_for_requests(request_rows: list[dict[str, Any]]) -> list[str]:
    fields = DETAIL_FIELD_ORDER.copy()
    has_team = any(normalize_text((item.get("request") or {}).get("team")) for item in request_rows)
    if has_team and "team" not in fields:
        fields.insert(2, "team")
    return fields


def field_label(field: str, template: dict[str, Any]) -> str:
    labels = template.get("field_labels") or {}
    if isinstance(labels, dict):
        value = normalize_text(labels.get(field))
        if value:
            return value
    return field


def html_paragraph(lines: list[str]) -> str:
    rendered: list[str] = []
    for line in lines:
        if not normalize_text(line):
            rendered.append("<div style='height:10px;'></div>")
        else:
            rendered.append(f"<p style='margin:0 0 8px 0;'>{escape(line)}</p>")
    return "".join(rendered)


def parse_request_date(sheet: Any) -> str:
    for row_idx in range(1, min(sheet.max_row, 4) + 1):
        for col_idx in range(1, min(sheet.max_column, 4) + 1):
            text = normalize_text(sheet.cell(row=row_idx, column=col_idx).value)
            if "신청일" not in text:
                continue
            match = re.search(r"(\d{4}[./-]\d{1,2}[./-]\d{1,2})", text)
            if match:
                return match.group(1).replace(".", "/").replace("-", "/")
    return ""


def build_context_from_requests(request_rows: list[dict[str, Any]], request_date: str) -> dict[str, str]:
    names = [normalize_text(item.get("request", {}).get("employee_name")) for item in request_rows]
    depts = [normalize_text(item.get("request", {}).get("department")) for item in request_rows]
    names = [item for item in names if item]
    depts = [item for item in depts if item]
    return {
        "request_date": request_date or datetime.now().strftime("%Y/%m/%d"),
        "request_count": str(len(request_rows)),
        "employee_name": names[0] if names else "",
        "department": depts[0] if depts else "",
        "employee_names": ", ".join(names[:8]) + (" 외" if len(names) > 8 else ""),
    }


def build_mail_html(context: dict[str, str], request_rows: list[dict[str, Any]], template: dict[str, Any]) -> str:
    intro_lines = [render_template_text(str(line), context) for line in template.get("intro_lines") or []]
    closing_lines = [render_template_text(str(line), context) for line in template.get("closing_lines") or []]

    summary_rows = [
        ("request_date", context.get("request_date", "")),
        ("request_count", context.get("request_count", "")),
        ("employee_names", context.get("employee_names", "")),
    ]
    summary_html: list[str] = []
    for field, value in summary_rows:
        if not normalize_text(value):
            continue
        label = field_label(field, template)
        summary_html.append(
            "<tr>"
            f"<th style='text-align:left;padding:8px 10px;background:#eff5ff;border:1px solid #c8d8f3;width:180px;'>{escape(label)}</th>"
            f"<td style='padding:8px 10px;border:1px solid #c8d8f3;'>{escape(value)}</td>"
            "</tr>"
        )

    detail_fields = detail_fields_for_requests(request_rows)
    detail_headers = "".join(
        f"<th style='text-align:left;padding:8px 10px;background:#eff5ff;border:1px solid #c8d8f3;white-space:nowrap;'>{escape(field_label(field, template))}</th>"
        for field in detail_fields
    )

    preview_rows: list[str] = []
    for item in request_rows[:20]:
        request = item.get("request") or {}
        cells = "".join(
            f"<td style='padding:7px 9px;border:1px solid #d6e1f1;vertical-align:top;'>{escape(display_cell(request.get(field)))}</td>"
            for field in detail_fields
        )
        preview_rows.append("<tr>" + cells + "</tr>")

    detail_title = "\uc2e0\uccad\uc11c \uc0c1\uc138"
    return f"""
<div style="font-family:'Noto Sans KR','Malgun Gothic',sans-serif;color:#21304d;font-size:10pt;line-height:1.6;">
  <div style="height:14px;background:linear-gradient(90deg,#1f4f99 0%,#4a8ee8 100%);border-radius:10px 10px 0 0;"></div>
  <div style="padding:24px 26px;border:1px solid #c8d8f3;border-top:none;background:#ffffff;">
    <div data-mail-section='intro'>{html_paragraph(intro_lines)}</div>
    <div data-mail-section='order-details'>
    <div style="height:10px;"></div>
    <table style="border-collapse:collapse;width:100%;max-width:900px;font-size:10pt;">
      {''.join(summary_html)}
    </table>
    <div style="height:18px;"></div>
    <p style="margin:0 0 10px 0;font-weight:700;">{detail_title}</p>
    <div style="overflow-x:auto;">
      <table style="border-collapse:collapse;width:100%;min-width:1200px;font-size:10pt;">
        <tr>{detail_headers}</tr>
        {''.join(preview_rows)}
      </table>
    </div>
    <div style="height:18px;"></div>
    </div>
    <div data-mail-section='closing'>{html_paragraph(closing_lines)}</div>
  </div>
  <div style="height:14px;background:linear-gradient(90deg,#4a8ee8 0%,#1f4f99 100%);border-radius:0 0 10px 10px;"></div>
</div>"""


def load_mapping() -> dict[str, Any]:
    raw = load_json(MAPPING_PATH, DEFAULT_MAPPING)
    field_aliases = raw.get("field_aliases") or {}
    normalized: dict[str, set[str]] = {}
    for field, aliases in field_aliases.items():
        values = {normalize_header(field)}
        for alias in aliases or []:
            norm = normalize_header(alias)
            if norm:
                values.add(norm)
        normalized[str(field)] = values
    raw.setdefault("minimum_header_matches", 4)
    raw.setdefault("required_header_fields", ["employee_name"])
    raw["_normalized_aliases"] = normalized
    return raw


def load_template() -> dict[str, Any]:
    return load_json(TEMPLATE_PATH, DEFAULT_TEMPLATE)


def discover_request_files() -> list[Path]:
    if not REQUEST_DIR.exists():
        REQUEST_DIR.mkdir(parents=True, exist_ok=True)
    files = [path for path in REQUEST_DIR.iterdir() if path.suffix.lower() in {".xlsx", ".xlsm"}]
    return sorted(files, key=lambda item: item.name.lower())


def load_state() -> dict[str, Any]:
    state = load_json(
        STATE_PATH,
        {
            "sent_hashes": {},
            "send_history": [],
            "fetched_message_uids": {},
            "import_history": [],
        },
    )
    if not isinstance(state, dict):
        return {
            "sent_hashes": {},
            "send_history": [],
            "fetched_message_uids": {},
            "import_history": [],
        }
    if not isinstance(state.get("sent_hashes"), dict):
        state["sent_hashes"] = {}
    if not isinstance(state.get("send_history"), list):
        state["send_history"] = []
    if not isinstance(state.get("fetched_message_uids"), dict):
        state["fetched_message_uids"] = {}
    if not isinstance(state.get("import_history"), list):
        state["import_history"] = []
    return state


def save_state(payload: dict[str, Any]) -> None:
    save_json(STATE_PATH, payload)


def choose_sheet_order(workbook: Any) -> list[Any]:
    named = []
    for name in workbook.sheetnames:
        sheet = workbook[name]
        if sheet.sheet_state != "visible":
            continue
        named.append(sheet)
    return named


def validate_header_match(matched: Mapping[str, int], mapping: Mapping[str, Any]) -> tuple[bool, str]:
    minimum = max(int(mapping.get("minimum_header_matches") or 4), 1)
    if len(matched) < minimum:
        return False, f"insufficient_header_matches:{len(matched)}/{minimum}"
    for field in mapping.get("required_header_fields") or ["employee_name"]:
        field_name = str(field)
        if field_name not in matched:
            return False, f"missing_header:{field_name}"
    return True, "accepted"


def find_best_header_row(sheet: Any, mapping: dict[str, Any]) -> tuple[int, dict[str, int]] | None:
    candidates = mapping.get("header_row_candidates") or [1, 2, 3, 4, 5]
    alias_map = mapping.get("_normalized_aliases") or {}
    best_row = 0
    best_match: dict[str, int] = {}
    best_score = -1

    for row_idx in candidates:
        matched: dict[str, int] = {}
        for col_idx in range(1, sheet.max_column + 1):
            header = normalize_header(sheet.cell(row=row_idx, column=col_idx).value)
            if not header:
                continue
            for field, aliases in alias_map.items():
                if field in matched:
                    continue
                if header in aliases:
                    matched[field] = col_idx
                    break
        score = len(matched)
        if score > best_score:
            best_row = int(row_idx)
            best_match = matched
            best_score = score
    accepted, _ = validate_header_match(best_match, mapping)
    if not accepted:
        return None
    return best_row, best_match


def extract_requests_from_workbook(path: Path, mapping: dict[str, Any]) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True)
    requests: list[dict[str, Any]] = []
    required_fields = [str(item) for item in mapping.get("required_fields") or []]
    normalized_fields = list((mapping.get("_normalized_aliases") or {}).keys())

    for sheet in choose_sheet_order(workbook):
        header_info = find_best_header_row(sheet, mapping)
        if not header_info:
            continue
        header_row, matched = header_info
        request_date = parse_request_date(sheet)
        for row_idx in range(header_row + 1, sheet.max_row + 1):
            values: dict[str, str] = {}
            nonempty = 0
            for field in normalized_fields:
                col_idx = matched.get(field)
                if not col_idx:
                    values[field] = ""
                    continue
                text = normalize_text(sheet.cell(row=row_idx, column=col_idx).value)
                values[field] = text
                if text:
                    nonempty += 1
            if nonempty == 0:
                continue
            missing = [field for field in required_fields if not normalize_text(values.get(field))]
            requests.append(
                {
                    "source_file": str(path),
                    "source_sheet": sheet.title,
                    "source_row": row_idx,
                    "request_date": request_date,
                    "request": values,
                    "missing_fields": missing,
                }
            )
    return requests


def split_items(raw: str) -> list[str]:
    if not normalize_text(raw):
        return []
    return [item.strip() for item in re.split(r"[;,\n]", str(raw)) if item.strip()]


def extract_sender_address(message: Mapping[str, Any]) -> str:
    from_obj = message.get("from")
    if isinstance(from_obj, dict):
        return normalize_text(from_obj.get("address") or from_obj.get("fullAddress"))
    sender = normalize_text(message.get("sender") or message.get("fromAddress"))
    match = re.search(r"<([^>]+)>", sender)
    if match:
        return normalize_text(match.group(1))
    return sender


def business_card_mailbox_config() -> dict[str, Any]:
    return {
        "monitor_interval_sec": max(int(os.getenv("BUSINESS_CARD_MONITOR_INTERVAL_SEC") or "60"), 10),
        "folder_name": normalize_text(os.getenv("BUSINESS_CARD_MAILBOX_FOLDER") or os.getenv("ARCHIVE_FROM_ADDRESS") or "INBOX"),
        "search_keyword": normalize_text(os.getenv("BUSINESS_CARD_MAILBOX_SEARCH") or ""),
        "take": max(int(os.getenv("BUSINESS_CARD_MAILBOX_TAKE") or "30"), 1),
        "keywords": [item.lower() for item in split_items(os.getenv("BUSINESS_CARD_MAILBOX_KEYWORDS") or "명함")],
        "allowed_senders": [item.lower() for item in split_items(os.getenv("BUSINESS_CARD_REQUEST_SENDERS") or "")],
    }


def message_uid(message: Mapping[str, Any]) -> str:
    email = normalize_text(message.get("email"))
    mail_id = normalize_text(message.get("id"))
    if email and mail_id:
        return f"{email}|{mail_id}"
    raw = "|".join([
        normalize_text(message.get("subject")),
        normalize_text(message.get("receivedDate")),
        extract_sender_address(message),
    ])
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()


def attachment_extension(attachment: Mapping[str, Any]) -> str:
    name = normalize_text(attachment.get("fileName"))
    ext = normalize_text(attachment.get("fileExt")).lower().lstrip(".")
    if ext:
        return ext
    return Path(name).suffix.lower().lstrip(".")


def is_excel_attachment(attachment: Mapping[str, Any]) -> bool:
    return attachment_extension(attachment) in {"xlsx", "xlsm"}


def classify_business_card_attachment(
    message: Mapping[str, Any], attachment: Mapping[str, Any], config: Mapping[str, Any]
) -> tuple[bool, str]:
    extension = attachment_extension(attachment)
    if extension == "xls":
        return False, "legacy_xls_not_supported"
    if extension not in {"xlsx", "xlsm"}:
        return False, "unsupported_attachment_type"
    allowed_senders = config.get("allowed_senders") or []
    sender_address = extract_sender_address(message).lower()
    if allowed_senders and sender_address not in allowed_senders:
        return False, "sender_not_allowed"
    subject = normalize_text(message.get("subject")).lower()
    keywords = config.get("keywords") or []
    if any(keyword and keyword in subject for keyword in keywords):
        return True, "accepted"
    return False, "keyword_not_matched"


def is_business_card_request_attachment(message: Mapping[str, Any], attachment: Mapping[str, Any], config: dict[str, Any]) -> bool:
    accepted, _ = classify_business_card_attachment(message, attachment, config)
    return accepted


def sanitize_filename(name: str) -> str:
    cleaned = re.sub(r'[\/:*?"<>|]+', '_', normalize_text(name))
    return cleaned or f"request_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"


def choose_download_path(file_name: str, payload: bytes) -> Path:
    REQUEST_DIR.mkdir(parents=True, exist_ok=True)
    safe_name = sanitize_filename(file_name)
    candidate = REQUEST_DIR / safe_name
    if not candidate.exists():
        return candidate
    try:
        if candidate.read_bytes() == payload:
            return candidate
    except Exception:
        pass
    stem = Path(safe_name).stem
    suffix = Path(safe_name).suffix or ".xlsx"
    short = hashlib.sha1(payload).hexdigest()[:8]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return REQUEST_DIR / f"{stem}_{timestamp}_{short}{suffix}"


def download_mail_attachment(client: Any, message: Mapping[str, Any], attachment: Mapping[str, Any]) -> bytes:
    email = normalize_text(message.get("email"))
    mail_id = normalize_text(message.get("id"))
    save_file_name = normalize_text(attachment.get("saveFileName"))
    if not email or not mail_id or not save_file_name:
        raise RuntimeError("첨부파일 다운로드 정보가 없어 파일을 받을 수 없습니다.")
    response = client.request(
        "GET",
        "/archive/download/attachment",
        params={"email": email, "id": mail_id, "saveFileName": save_file_name},
        accept="*/*",
        referer_path="/archive",
    )
    content_type = normalize_text(response.headers.get("Content-Type"))
    if "text/html" in content_type.lower() and "error page" in response.text.lower():
        raise RuntimeError("첨부파일 다운로드 요청이 실패했습니다.")
    if not response.body:
        raise RuntimeError("첨부파일 내용이 비어 있습니다.")
    return response.body


def fetch_requests_from_mail(max_messages: int | None = None) -> dict[str, Any]:
    state = load_state()
    fetched_message_uids = state.get("fetched_message_uids") or {}
    config = business_card_mailbox_config()
    take = max_messages or int(config.get("take") or 30)
    module = load_archive_module()
    session = module.open_mailer_session(run_id="business_card_fetch")
    results: list[dict[str, Any]] = []
    scan_count = 0
    try:
        client = session["client"]
        folders = client.fetch_mailboxes()
        folder = client.find_folder(folders, str(config.get("folder_name") or "INBOX"))
        messages = client.list_messages(folder, take=take, keyword=str(config.get("search_keyword") or ""))
        scan_count = len(messages)
        for message in messages:
            uid = message_uid(message)
            attachments = message.get("attachmentsWithoutCid") or message.get("attaches") or message.get("attachments") or []
            matching = [
                attachment for attachment in attachments
                if isinstance(attachment, dict) and is_business_card_request_attachment(message, attachment, config)
            ]
            if not matching:
                continue
            if uid in fetched_message_uids:
                results.append(
                    {
                        "uid": uid,
                        "subject": normalize_text(message.get("subject")),
                        "sender": extract_sender_address(message),
                        "status": "skipped",
                        "reason": "already_imported",
                        "saved_files": fetched_message_uids.get(uid, {}).get("saved_files", []),
                    }
                )
                continue
            saved_files: list[str] = []
            for attachment in matching:
                payload = download_mail_attachment(client, message, attachment)
                save_path = choose_download_path(str(attachment.get("fileName") or "request.xlsx"), payload)
                if not save_path.exists():
                    save_path.write_bytes(payload)
                saved_files.append(str(save_path))
            fetched_message_uids[uid] = {
                "fetched_at": datetime.now().isoformat(timespec="seconds"),
                "subject": normalize_text(message.get("subject")),
                "sender": extract_sender_address(message),
                "saved_files": saved_files,
            }
            results.append(
                {
                    "uid": uid,
                    "subject": normalize_text(message.get("subject")),
                    "sender": extract_sender_address(message),
                    "status": "imported",
                    "saved_files": saved_files,
                }
            )
    finally:
        module.close_mailer_session(session, run_id="business_card_fetch")

    result_payload = {
        "fetched_at": datetime.now().isoformat(timespec="seconds"),
        "mail_scan_count": scan_count,
        "imported_count": sum(1 for item in results if item.get("status") == "imported"),
        "skipped_count": sum(1 for item in results if item.get("status") == "skipped"),
        "results": results,
    }
    state["fetched_message_uids"] = fetched_message_uids
    history = state.get("import_history") or []
    history.insert(
        0,
        {
            "fetched_at": result_payload["fetched_at"],
            "mail_scan_count": result_payload["mail_scan_count"],
            "imported_count": result_payload["imported_count"],
            "skipped_count": result_payload["skipped_count"],
            "note": "메일함에서 신청서를 자동으로 가져옴",
        },
    )
    state["import_history"] = history[:50]
    save_state(state)
    save_json(MAIL_FETCH_RESULT_PATH, result_payload)
    write_dashboard()
    return result_payload


def build_drafts(include_sent: bool = False) -> dict[str, Any]:
    mapping = load_mapping()
    template = load_template()
    state = load_state()
    sent_hashes = state.get("sent_hashes") or {}
    vendor_to, vendor_cc = get_vendor_recipient(template)

    request_files = discover_request_files()
    drafts: list[DraftRecord] = []
    skipped_sent: list[str] = []
    parse_errors: list[dict[str, str]] = []
    rejected_files: list[dict[str, str]] = []

    for path in request_files:
        source_hash = file_sha1(path)
        if not include_sent and source_hash in sent_hashes:
            skipped_sent.append(path.name)
            continue
        try:
            extracted = extract_requests_from_workbook(path, mapping)
        except Exception as exc:
            parse_errors.append({"file": path.name, "error": str(exc)})
            continue
        if not extracted:
            rejected_files.append({"file": path.name, "reason": "명함 신청서 핵심 헤더 기준을 충족하지 않습니다."})
            continue

        request_date = next((normalize_text(item.get("request_date")) for item in extracted if normalize_text(item.get("request_date"))), "")
        context = build_context_from_requests(extracted, request_date=request_date)
        subject = render_template_text(str(template.get("subject_template") or ""), context)
        missing: list[str] = []
        for item in extracted:
            for field in item.get("missing_fields") or []:
                if field not in missing:
                    missing.append(field)
        if not vendor_to:
            missing.append("vendor_to")
        status = "ready" if not missing else "pending"
        draft_id = hashlib.sha1(f"{source_hash}|{extracted[0]['source_sheet']}".encode("utf-8")).hexdigest()[:12]
        drafts.append(
            DraftRecord(
                draft_id=draft_id,
                source_file=str(path),
                source_sheet=str(extracted[0]["source_sheet"]),
                source_hash=source_hash,
                request_date=context["request_date"],
                request_count=len(extracted),
                status=status,
                missing_fields=missing,
                subject=subject,
                vendor_to=vendor_to,
                vendor_cc=vendor_cc,
                html_body=build_mail_html(context, extracted, template),
                attachment_paths=[str(path)] if bool(template.get("attach_original_excel", True)) else [],
                requests=extracted,
            )
        )

    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "request_file_count": len(request_files),
        "draft_count": len(drafts),
        "ready_count": sum(1 for item in drafts if item.status == "ready"),
        "pending_count": sum(1 for item in drafts if item.status == "pending"),
        "skipped_sent_files": skipped_sent,
        "parse_errors": parse_errors,
        "rejected_files": rejected_files,
        "drafts": [asdict(item) for item in drafts],
    }
    save_json(DRAFTS_PATH, payload)
    PREVIEW_PATH.write_text(render_preview_html(payload, template), encoding="utf-8")
    write_dashboard(payload=payload, template=template)
    return payload


def render_preview_html(payload: dict[str, Any], template: dict[str, Any]) -> str:
    ready_badge = "<span class='badge ready'>\ubc1c\uc1a1 \uac00\ub2a5</span>"
    pending_badge = "<span class='badge pending'>\ud655\uc778 \ud544\uc694</span>"
    untitled = "(\uc81c\ubaa9 \uc5c6\uc74c)"
    file_label = "\ud30c\uc77c"
    sheet_label = "\uc2dc\ud2b8"
    to_label = "\uc218\uc2e0"
    cc_label = "\ucc38\uc870"
    issue_title = "\ud655\uc778 \ud544\uc694"
    page_title = "\uba85\ud568 \ubc1c\uc8fc \uba54\uc77c \ubbf8\ub9ac\ubcf4\uae30"
    page_desc = "\uc2e0\uccad\uc11c \uc6d0\ubcf8 \ud30c\uc77c\uc744 \uc5c5\uccb4\uc5d0 \uc804\ub2ec\ud558\uae30 \uc704\ud55c \ubc1c\uc8fc \uba54\uc77c \ucd08\uc548\uc785\ub2c8\ub2e4."
    stat_request_files = "\uc2e0\uccad \ud30c\uc77c"
    stat_drafts = "\uc0dd\uc131\ub41c \ucd08\uc548"
    stat_ready = "\ubc1c\uc1a1 \uac00\ub2a5"
    stat_pending = "\ud655\uc778 \ud544\uc694"
    panel_result = "\ud30c\uc77c \ucc98\ub9ac \uacb0\uacfc"
    generated_label = "\uc0dd\uc131 \uc2dc\uac01"
    skipped_label = "\uc774\ubbf8 \ubc1c\uc1a1\ub418\uc5b4 \uc774\ubc88\uc5d0 \uc81c\uc678\ub41c \ud30c\uc77c"
    parse_error_label = "\ud30c\uc2f1 \uc624\ub958"
    empty_text = "\uc0dd\uc131\ub41c \ucd08\uc548\uc774 \uc5c6\uc2b5\ub2c8\ub2e4."
    row_label = "\ud589"

    cards: list[str] = []
    for item in payload.get("drafts") or []:
        labels = [
            f"<tr><th>{escape(field_label('request_date', template))}</th><td>{escape(str(item.get('request_date') or '-'))}</td></tr>",
            f"<tr><th>{escape(field_label('request_count', template))}</th><td>{escape(str(item.get('request_count') or 0))}</td></tr>",
        ]
        missing_fields = [
            f"<li>{escape(field_label(field, template))}</li>"
            for field in item.get("missing_fields") or []
        ]
        badges = ready_badge if item.get("status") == "ready" else pending_badge

        request_rows = item.get("requests") or []
        detail_fields = detail_fields_for_requests(request_rows)
        header_html = f"<tr><th>{row_label}</th>" + "".join(
            f"<th>{escape(field_label(field, template))}</th>"
            for field in detail_fields
        ) + "</tr>"

        preview_rows: list[str] = []
        for row in request_rows:
            request = row.get("request") or {}
            cells = [f"<td>{escape(str(row.get('source_row') or ''))}</td>"]
            cells.extend(
                f"<td>{escape(display_cell(request.get(field)))}</td>"
                for field in detail_fields
            )
            preview_rows.append("<tr>" + "".join(cells) + "</tr>")

        cards.append(
            "<details class='draft-card' open>"
            f"<summary>{badges} {escape(item.get('subject') or untitled)}</summary>"
            f"<div class='meta'>{file_label}: {escape(Path(item.get('source_file') or '').name)} / {sheet_label}: {escape(str(item.get('source_sheet') or ''))}</div>"
            f"<div class='meta'>{to_label}: {escape(item.get('vendor_to') or '(\ubbf8\uc124\uc815)')} / {cc_label}: {escape(item.get('vendor_cc') or '-')}</div>"
            "<table class='field-table'>"
            f"{''.join(labels)}"
            "</table>"
            "<div class='table-scroll'><table class='field-table detail-table'>"
            f"{header_html}"
            f"{''.join(preview_rows)}"
            "</table></div>"
            + (
                f"<div class='issue-box'><strong>{issue_title}</strong><ul>"
                + "".join(missing_fields)
                + "</ul></div>"
                if missing_fields
                else ""
            )
            + "</details>"
        )

    errors = [
        f"<li>{escape(item.get('file') or '')}: {escape(item.get('error') or '')}</li>"
        for item in payload.get("parse_errors") or []
    ]
    skipped = [
        f"<li>{escape(name)}</li>"
        for name in payload.get("skipped_sent_files") or []
    ]

    return f"""<!doctype html>
<html lang=\"ko\">
<head>
  <meta charset=\"utf-8\">
  <title>{page_title}</title>
  <style>
    body {{
      font-family: 'Noto Sans KR','Malgun Gothic',sans-serif;
      background: #f4f7fb;
      color: #21304d;
      margin: 0;
      padding: 28px;
    }}
    .wrap {{
      max-width: 1200px;
      margin: 0 auto;
    }}
    .hero {{
      background: linear-gradient(135deg, #1f4f99 0%, #5ea4ff 100%);
      color: white;
      padding: 24px 28px;
      border-radius: 18px;
      box-shadow: 0 14px 32px rgba(31, 79, 153, 0.18);
    }}
    .stats {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(140px, 1fr));
      gap: 14px;
      margin: 24px 0;
    }}
    .card, .panel, .draft-card {{
      background: #ffffff;
      border-radius: 16px;
      box-shadow: 0 10px 24px rgba(22, 40, 78, 0.08);
      border: 1px solid #d9e4f5;
    }}
    .card {{
      padding: 16px 18px;
    }}
    .card strong {{
      display: block;
      font-size: 28px;
      color: #1f4f99;
      margin-bottom: 4px;
    }}
    .panel {{
      padding: 18px 20px;
      margin-bottom: 18px;
    }}
    .draft-card {{
      margin-bottom: 14px;
      overflow: hidden;
    }}
    .draft-card summary {{
      cursor: pointer;
      list-style: none;
      padding: 16px 18px;
      font-weight: 700;
      background: #f9fbff;
      border-bottom: 1px solid #d9e4f5;
    }}
    .draft-card summary::-webkit-details-marker {{
      display: none;
    }}
    .meta {{
      padding: 10px 18px 0 18px;
      color: #4a5875;
      font-size: 14px;
    }}
    .table-scroll {{
      overflow-x: auto;
      margin: 0 18px 18px 18px;
    }}
    .field-table {{
      width: calc(100% - 36px);
      margin: 14px 18px 18px 18px;
      border-collapse: collapse;
    }}
    .detail-table {{
      width: 100%;
      min-width: 1300px;
      margin: 14px 0 0 0;
    }}
    .field-table th, .field-table td {{
      border: 1px solid #d6e1f1;
      padding: 8px 10px;
      text-align: left;
      vertical-align: top;
      font-size: 14px;
    }}
    .field-table th {{
      background: #eff5ff;
      white-space: nowrap;
    }}
    .badge {{ display:inline-block; padding:5px 10px; border-radius:999px; font-size:12px; font-weight:700; margin-right:8px; }}
    .badge.ready {{ background:#e5f7ea; color:#1f7b43; }}
    .badge.pending {{ background:#fff1de; color:#b46800; }}
    .issue-box {{ margin:0 18px 18px 18px; background:#fff8ea; border:1px solid #f4d48a; border-radius:12px; padding:14px 16px; }}
    ul {{ margin:8px 0 0 18px; }}
  </style>
</head>
<body>
  <div class=\"wrap\">
    <section class=\"hero\">
      <h1>{page_title}</h1>
      <p>{page_desc}</p>
    </section>

    <section class=\"stats\">
      <div class=\"card\"><strong>{payload.get('request_file_count', 0)}</strong>{stat_request_files}</div>
      <div class=\"card\"><strong>{payload.get('draft_count', 0)}</strong>{stat_drafts}</div>
      <div class=\"card\"><strong>{payload.get('ready_count', 0)}</strong>{stat_ready}</div>
      <div class=\"card\"><strong>{payload.get('pending_count', 0)}</strong>{stat_pending}</div>
    </section>

    <section class=\"panel\">
      <h2>{panel_result}</h2>
      <p>{generated_label}: {escape(str(payload.get('generated_at') or '-'))}</p>
      <p>{skipped_label}: {len(payload.get('skipped_sent_files') or [])}</p>
      <p>{parse_error_label}: {len(payload.get('parse_errors') or [])}</p>
      {('<ul>' + ''.join(skipped) + '</ul>') if skipped else ''}
      {('<ul>' + ''.join(errors) + '</ul>') if errors else ''}
    </section>

    {''.join(cards) if cards else f'<section class="panel">{empty_text}</section>'}
  </div>
</body>
</html>"""


def load_drafts_payload() -> dict[str, Any]:
    return load_json(DRAFTS_PATH, {"drafts": []})


def load_send_result() -> dict[str, Any]:
    return load_json(SEND_RESULT_PATH, {})


def load_fetch_result() -> dict[str, Any]:
    return load_json(MAIL_FETCH_RESULT_PATH, {})


def render_dashboard_html(payload: dict[str, Any], template: dict[str, Any], state: dict[str, Any], send_result: dict[str, Any], fetch_result: dict[str, Any], send_mode: str = "manual") -> str:
    """Compatibility wrapper for callers that import the renderer from this module."""
    return render_dashboard_page(payload, template, state, send_result, fetch_result, send_mode=send_mode)

def write_dashboard(payload: dict[str, Any] | None = None, template: dict[str, Any] | None = None) -> None:
    current_payload = payload or load_drafts_payload()
    current_template = template or load_template()
    state = load_state()
    send_result = load_send_result()
    fetch_result = load_fetch_result()
    DASHBOARD_PATH.write_text(
        render_dashboard_html(
            current_payload,
            current_template,
            state,
            send_result,
            fetch_result,
            send_mode=load_automation_settings(AUTOMATION_SETTINGS_PATH)["send_mode"],
        ),
        encoding="utf-8",
    )


def load_archive_module() -> Any:
    global _ARCHIVE_MODULE
    if _ARCHIVE_MODULE is not None:
        return _ARCHIVE_MODULE
    archive_path = PROJECT_DIR.parent / "3. SOA_fup_sales" / "archive_mailer_api.py"
    if not archive_path.exists():
        raise FileNotFoundError(f"기존 메일 발송 모듈을 찾을 수 없습니다: {archive_path}")
    spec = importlib.util.spec_from_file_location("shared_archive_mailer_api", archive_path)
    if spec is None or spec.loader is None:
        raise RuntimeError("기존 메일 발송 모듈을 불러올 수 없습니다.")
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    try:
        spec.loader.exec_module(module)
    except Exception:
        sys.modules.pop(spec.name, None)
        raise
    _ARCHIVE_MODULE = module
    return module


def send_ready_drafts(
    approve_send: bool,
    archive_module: Any | None = None,
    payload: Mapping[str, Any] | None = None,
    operation_locked: bool = False,
) -> dict[str, Any]:
    if not approve_send:
        raise RuntimeError("실제 발송은 --approve-send 옵션이 있을 때만 가능합니다.")
    module = archive_module or load_archive_module()
    result_payload = send_ready_drafts_safely(
        payload=dict(payload) if payload is not None else load_drafts_payload(),
        state=None,
        archive_module=module,
        save_state_callback=save_state,
        save_result_callback=lambda payload: save_json(SEND_RESULT_PATH, payload),
        lock_path=OPERATION_LOCK_PATH,
        file_hasher=file_sha1,
        state_loader=load_state,
        operation_locked=operation_locked,
    )
    refreshed_payload = build_drafts(include_sent=False)
    write_dashboard(payload=refreshed_payload)
    return result_payload


def cmd_build(args: argparse.Namespace) -> int:
    payload = build_drafts(include_sent=args.include_sent)
    print(f"메일 초안: {payload.get('draft_count', 0)}건")
    print(f"발송 가능: {payload.get('ready_count', 0)}건 / 확인 필요: {payload.get('pending_count', 0)}건")
    print(f"미리보기: {PREVIEW_PATH}")
    print(f"대시보드: {DASHBOARD_PATH}")
    if args.prompt_send and payload.get("ready_count", 0) > 0 and payload.get("pending_count", 0) == 0:
        answer = input("실제 업체로 메일을 발송하시겠습니까? [y/N]: ").strip().lower()
        if answer == "y":
            send_payload = send_ready_drafts(approve_send=True)
            print(f"발송 결과: 성공 {send_payload.get('success_count', 0)}건 / 실패 {send_payload.get('fail_count', 0)}건")
    return 0


def cmd_fetch(args: argparse.Namespace) -> int:
    payload = fetch_requests_from_mail(max_messages=args.take)
    print(f"메일 확인: {payload.get('mail_scan_count', 0)}건")
    print(f"새 신청서: {payload.get('imported_count', 0)}건 / 제외: {payload.get('skipped_count', 0)}건")
    print(f"수집 결과: {MAIL_FETCH_RESULT_PATH}")
    print(f"대시보드: {DASHBOARD_PATH}")
    return 0

def _unlocked_run_sync_cycle(take: int = 0, include_sent: bool = False) -> dict[str, Any]:
    fetch_payload = fetch_requests_from_mail(max_messages=take)
    draft_payload = build_drafts(include_sent=include_sent)
    return {"fetch": fetch_payload, "drafts": draft_payload}


def run_sync_cycle(take: int = 0, include_sent: bool = False) -> dict[str, Any]:
    with OperationLock(OPERATION_LOCK_PATH):
        return _unlocked_run_sync_cycle(take=take, include_sent=include_sent)


def run_locked_automation_cycle(
    settings_path: Path = AUTOMATION_SETTINGS_PATH,
) -> dict[str, Any]:
    with OperationLock(OPERATION_LOCK_PATH):
        return run_automation_cycle(
            sync_cycle=lambda: _unlocked_run_sync_cycle(),
            send_payload=lambda payload: send_ready_drafts(
                approve_send=True,
                payload=payload,
                operation_locked=True,
            ),
            settings_path=settings_path,
        )


def cmd_sync(args: argparse.Namespace) -> int:
    result = run_sync_cycle(take=args.take, include_sent=args.include_sent)
    fetch_payload = result["fetch"]
    payload = result["drafts"]
    print(f"메일 확인: {fetch_payload.get('mail_scan_count', 0)}건")
    print(f"새 신청서: {fetch_payload.get('imported_count', 0)}건 / 제외: {fetch_payload.get('skipped_count', 0)}건")
    print(f"메일 초안: {payload.get('draft_count', 0)}건")
    print(f"발송 가능: {payload.get('ready_count', 0)}건 / 확인 필요: {payload.get('pending_count', 0)}건")
    print(f"미리보기: {PREVIEW_PATH}")
    print(f"대시보드: {DASHBOARD_PATH}")
    return 0


def cmd_monitor(args: argparse.Namespace) -> int:
    interval = max(int(args.interval or business_card_mailbox_config().get("monitor_interval_sec") or 60), 10)
    print(f"자동 확인 간격: {interval}초")
    print("종료하려면 Ctrl+C를 누르세요.")
    while True:
        started_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            result = run_sync_cycle(take=args.take, include_sent=args.include_sent)
            print(f"[{started_at}] 메일 {result['fetch'].get('mail_scan_count', 0)}건 확인 / 새 신청서 {result['fetch'].get('imported_count', 0)}건 / 발송 가능 {result['drafts'].get('ready_count', 0)}건")
        except KeyboardInterrupt:
            print("자동 확인을 종료합니다.")
            return 0
        except Exception as exc:
            print(f"[{started_at}] 자동 확인 실패: {exc}")
        try:
            time.sleep(interval)
        except KeyboardInterrupt:
            print("자동 확인을 종료합니다.")
            return 0


def cmd_send(args: argparse.Namespace) -> int:
    try:
        payload = send_ready_drafts(approve_send=args.approve_send)
    except Exception as exc:
        failure_payload = {
            "sent_at": datetime.now().isoformat(timespec="seconds"),
            "total_count": 0,
            "success_count": 0,
            "fail_count": 1,
            "results": [{"ok": False, "subject": "명함 발주 실패", "message": str(exc), "sent_at": datetime.now().isoformat(timespec="seconds")}],
        }
        state = load_state()
        history = state.get("send_history") or []
        history.insert(0, {"sent_at": failure_payload["sent_at"], "success_count": 0, "fail_count": 1, "total_count": 0, "note": "발주 메일 전송 실패"})
        state["send_history"] = history[:50]
        save_state(state)
        save_json(SEND_RESULT_PATH, failure_payload)
        write_dashboard()
        print(f"발송 실패: {exc}")
        print(f"발송 결과: {SEND_RESULT_PATH}")
        print(f"대시보드: {DASHBOARD_PATH}")
        return 1
    print(f"발송 결과: 성공 {payload.get('success_count', 0)}건 / 실패 {payload.get('fail_count', 0)}건")
    print(f"발송 결과: {SEND_RESULT_PATH}")
    print(f"대시보드: {DASHBOARD_PATH}")
    return 0


def cmd_dashboard(args: argparse.Namespace) -> int:
    del args
    write_dashboard()
    print(f"대시보드 생성 완료: {DASHBOARD_PATH}")
    return 0

def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Business Card Order Mailer")
    sub = parser.add_subparsers(dest="command", required=True)

    fetch = sub.add_parser("fetch", help="메일함에서 명함 신청 첨부파일을 가져옵니다.")
    fetch.add_argument("--take", type=int, default=0, help="확인할 최근 메일 수입니다. 기본값은 환경설정을 사용합니다.")

    build = sub.add_parser("build", help="신청서를 분석해 발주 메일 초안을 만듭니다.")
    build.add_argument("--include-sent", action="store_true", help="이미 발송한 신청서도 포함합니다.")
    build.add_argument("--prompt-send", action="store_true", help="초안 생성 후 터미널에서 발송 승인을 받습니다.")

    sync = sub.add_parser("sync", help="메일 수집과 초안 생성을 한 번 실행합니다.")
    sync.add_argument("--take", type=int, default=0, help="확인할 최근 메일 수입니다.")
    sync.add_argument("--include-sent", action="store_true", help="이미 발송한 신청서도 포함합니다.")

    monitor = sub.add_parser("monitor", help="메일함을 주기적으로 확인하고 초안을 갱신합니다.")
    monitor.add_argument("--take", type=int, default=0, help="확인할 최근 메일 수입니다.")
    monitor.add_argument("--include-sent", action="store_true", help="이미 발송한 신청서도 포함합니다.")
    monitor.add_argument("--interval", type=int, default=0, help="메일 확인 간격(초)입니다.")

    send = sub.add_parser("send", help="승인된 메일 초안을 실제로 발송합니다.")
    send.add_argument("--approve-send", action="store_true", help="실제 발송을 명시적으로 승인합니다.")

    sub.add_parser("dashboard", help="수집·초안·발송 이력 대시보드를 다시 생성합니다.")
    return parser

def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    handlers = {
        'fetch': cmd_fetch,
        'build': cmd_build,
        'sync': cmd_sync,
        'monitor': cmd_monitor,
        'send': cmd_send,
        'dashboard': cmd_dashboard,
    }
    return handlers[args.command](args)


if __name__ == '__main__':
    raise SystemExit(main())











